
import os
from tkinter import *
import random
from tkinter.font import BOLD
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.workbook import workbook


class MainWindow():

    
    def __init__(self, mainwidget):

        self.start_gui()

    setting_status = TRUE

    title = ''
    contest = ''

    def start_gui(self):

        self.frame_start = Frame(root, bg= 'gray')
        self.frame_start.pack(expand= YES, fill= BOTH)

        self.img_logo = PhotoImage(file= 'ShSolutions.png')

        self.lbl_logo = Label(self.frame_start, bg= 'gray', image= self.img_logo)
        self.lbl_logo.pack(pady= 150, anchor= CENTER, fill= BOTH)

        self.frame_start.after(2000, self.main_gui, self.frame_start)

    def main_gui(self, frame):

        try:
            frame.destroy()
        except:
            pass

        self.frame_main = Frame(root, bg= 'gray')
        self.frame_main.pack(expand= YES, fill= BOTH)

        self.lbl_main = Label(self.frame_main, bg= 'gray', fg= '#FF9933', font= ('B Nazanin', 20, BOLD), text= f'ابزار قرعه کشی {self.title}')
        self.lbl_main.pack(ipadx= 5, ipady= 20, padx= 10, pady= (30, 15), fill= X)

        self.btn_open_file = Button(self.frame_main, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 16, BOLD), text= 'انتخاب فایل',
                                    command= lambda : self.open_file())
        self.btn_open_file.pack(padx= 30, ipady= 10, pady= 20, fill= X)
        if self.setting_status :
            self.btn_open_file.config(state= DISABLED)
        else :
            self.btn_open_file.config(state= NORMAL)

        self.btn_manual = Button(self.frame_main, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 16, BOLD), text= 'راهنمای استفاده')
        self.btn_manual.pack(padx= 30, ipady= 10, pady= 20, fill= X)
        self.btn_manual.bind('<Button-1>', self.manual_gui)

        
        self.btn_setting = Button(self.frame_main, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 16, BOLD), text= 'تنظیمات')
        self.btn_setting.pack(padx= 30, ipady= 10, pady= 20, fill= X)
        self.btn_setting.bind('<Button-1>', self.setting_gui)

        self.btn_exit = Button(self.frame_main, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 16, BOLD), text= 'خروج از برنامه')
        self.btn_exit.pack(padx= 30, ipady= 10, pady= 20, fill= X)
        self.btn_exit.bind('<Button-1>', self.exit_app)

        self.lbl_version = Label(self.frame_main, bg= 'gray', font= ('Helvetica', 8), text= 'Software Version : 1.0')
        self.lbl_version.pack(pady= (5, 10), side= BOTTOM, fill= X)

    def manual_gui(self, event):

        self.frame_main.destroy()

        self.frame_manual = Frame(root, bg= 'gray')
        self.frame_manual.pack(expand= YES, fill= BOTH)

        self.lbl_manual_title = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 17, BOLD), anchor= E, text= ': راهنمای استفاده')
        self.lbl_manual_title.pack(ipadx= 5, padx= 20, pady= (25, 10), fill= X)

        self.list_manual_text = ['.ابتدا لیست مورد نظر خود را در فایل اکسل قرار دهید -',
                                 '.از صفحه اصلی برنامه وارد بخش تنظیمات شوید -',
                                 'نام شرکت یا مجموعه خود را جهت نمایش در عنوان -\n.اصلی برنامه وارد کنید   ',
                                 '    .برای قرعه کشی خود عنوانی را انتخاب کنید -',
                                 'ستونی که قرعه کشی باید در بین آیتم های آن انجام -\n:شود را مانند مثال زیر وارد کنید   \n« B یا « عدد 2 برای ستون « A عدد 1 برای ستون »   ',
                                 'ردیفی که اولین آیتم مورد نظرتان در آن قرار دارد را -\nوارد کنید، یه عنوان مثال : اگر در ستون های لیستی   \nدر ردیف 1 عنوان قرار گرفته باشد، آیتم های لیست    \nاز ردیف 2 شروع می شوند و شما عدد 2 را در بخش    \n.ردیف وارد کنید    ',
                                 'پس از ذخیره تنظیمات در صفحه اصلی برنامه، دکمه -\n.انتخاب فایل » فعال میشود »    '
                                ]

        self.lbl_manual_text_1 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[0])
        self.lbl_manual_text_1.pack(padx= 20, fill= X)
        self.lbl_manual_text_2 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[1])
        self.lbl_manual_text_2.pack(padx= 20, fill= X)
        self.lbl_manual_text_3 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[2])
        self.lbl_manual_text_3.pack(padx= 20, fill= X)
        self.lbl_manual_text_4 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[3])
        self.lbl_manual_text_4.pack(padx= 20, fill= X)
        self.lbl_manual_text_5 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[4])
        self.lbl_manual_text_5.pack(padx= 20, fill= X)
        self.lbl_manual_text_6 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[5])
        self.lbl_manual_text_6.pack(padx= 20, fill= X)
        self.lbl_manual_text_6 = Label(self.frame_manual, bg= 'gray', fg= 'white', font= ('B Nazanin', 13, BOLD), anchor= E, justify= RIGHT, text= self.list_manual_text[6])
        self.lbl_manual_text_6.pack(padx= 20, fill= X)

        self.btn_back = Button(self.frame_manual, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'بازگشت', justify= CENTER,
                                command= lambda : self.main_gui(self.frame_manual))
        self.btn_back.pack(pady= 25, padx= 70, fill= X)

    def setting_gui(self, event):

        self.frame_main.destroy()

        self.frame_setting = Frame(root, bg= 'gray')
        self.frame_setting.pack(expand= YES, fill= BOTH)

        self.lbl_title = Label(self.frame_setting, bg= 'gray', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'نام مجموعه خود را وارد کنید', justify= CENTER)
        self.lbl_title.grid(row= 1, column= 0, columnspan= 2, pady= (40,10), padx= 70, sticky= EW)

        self.ent_title = Entry(self.frame_setting, bg= 'white', font= ('B Nazanin', 14, BOLD), justify= CENTER)
        self.ent_title.grid(row= 2, column= 0, columnspan= 2, pady= (10,20), padx= 70, sticky= EW)

        self.lbl_contest = Label(self.frame_setting, bg= 'gray', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'عنوان قرعه کشی را انتخاب کنید', justify= CENTER)
        self.lbl_contest.grid(row= 3, column= 0, columnspan= 2, pady= (20,10), padx= 70, sticky= EW)

        self.ent_contest = Entry(self.frame_setting, bg= 'white', font= ('B Nazanin', 14, BOLD), justify= CENTER)
        self.ent_contest.grid(row= 4, column= 0, columnspan= 2, pady= (10,20), padx= 70, sticky= EW)

        self.lbl_column = Label(self.frame_setting, bg= 'gray', fg= 'white', font= ('B Nazanin', 14, BOLD), text= ': ستون', justify= CENTER)
        self.lbl_column.grid(row= 5, column= 1, pady= (35,20), padx= (5, 120))

        self.ent_column = Entry(self.frame_setting, bg= 'white', font= ('B Nazanin', 14, BOLD), width= 8, justify= CENTER)
        self.ent_column.grid(row= 5, column= 0, pady= (35,20), padx= (120, 5))

        self.lbl_row = Label(self.frame_setting, bg= 'gray', fg= 'white', font= ('B Nazanin', 14, BOLD), text= ': ردیف', justify= CENTER)
        self.lbl_row.grid(row= 6, column= 1, pady= (10,30), padx= (5, 120))

        self.ent_row = Entry(self.frame_setting, bg= 'white', font= ('B Nazanin', 14, BOLD), width= 8, justify= CENTER)
        self.ent_row.grid(row= 6, column= 0, pady= (10,30), padx= (120, 5))

        self.btn_save_setting = Button(self.frame_setting, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'ذخیره کردن تنظیمات', justify= CENTER)
        self.btn_save_setting.grid(row= 7, column= 0, columnspan= 2, pady= (20,10), padx= 70, sticky= EW)
        self.btn_save_setting.bind('<Button-1>', self.save_setting)

        self.btn_back = Button(self.frame_setting, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'بازگشت', justify= CENTER,
                                command= lambda : self.main_gui(self.frame_setting))
        self.btn_back.grid(row= 8, column= 0, columnspan= 2, pady= (20,10), padx= 70, sticky= EW)
    
    def contest_gui(self):

        self.frame_main.destroy()

        self.winners_list = []
        self.index = IntVar

        self.frame_contest = Frame(root, bg= 'gray')
        self.frame_contest.pack(expand= YES, fill= BOTH)

        self.btn_contest = Button(self.frame_contest, bg= 'white', activebackground= '#ff5a00', font= ('B Nazanin', 14, BOLD), text= 'انجام قرعه کشی', justify= CENTER)
        self.btn_contest.pack(pady= (45, 15), padx= 70, fill= X)
        self.btn_contest.bind('<Button-1>', self.select_winner)

        self.lbl_winners = Label(self.frame_contest, bg= 'gray', fg= '#FF9933', font= ('B Nazanin', 16, BOLD), text= f'برندگان مسابقه {self.contest}', justify= CENTER)
        self.lbl_winners.pack(pady= (15, 10))

        self.frame_lbox = Frame(self.frame_contest)
        self.frame_lbox.pack(padx= 70, fill= BOTH)

        self.scrollbar = Scrollbar(self.frame_lbox, orient= VERTICAL)
        self.scrollbar.pack(side= RIGHT, fill= Y)
        
        self.lbox_winners = Listbox(self.frame_lbox, bg= 'white', font= ('Helvetica', 12, BOLD), height= 15)
        self.lbox_winners.pack(fill= BOTH)
        self.lbox_winners.config(yscrollcommand= self.scrollbar.set)
        self.scrollbar.config(command= self.lbox_winners.yview)

        self.btn_save = Button(self.frame_contest, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'ذخیره کردن اسامی برندگان', justify= CENTER)
        self.btn_save.pack(pady= (20,10), padx= 70, fill= X)
        self.btn_save.bind('<Button-1>', self.save_winners_list)

        self.btn_back = Button(self.frame_contest, bg= '#FF9933', fg= 'white', font= ('B Nazanin', 14, BOLD), text= 'بازگشت', justify= CENTER,
                                command= lambda : self.main_gui(self.frame_contest))
        self.btn_back.pack(pady= 10, padx= 70, fill= X)


    def save_setting(self, event):
        
        error_message = []
        
        try :
            column = int(self.ent_column.get())
            if column <= 0 :
                error_message.append('! برای ستون مقدار عددی بزرگتر از 0 وارد کنید')
                self.ent_column.delete(0, END)

        except :
            error_message.append('! برای ستون مقدار عددی (1, 2, 3 و ...) وارد کنید')
            self.ent_column.delete(0, END)
        try :
            row = int(self.ent_row.get())
            if row <= 0 :
                error_message.append('! برای ردیف مقدار عددی بزرگتر از 0 وارد کنید')
                self.ent_row.delete(0, END)
        except :
            error_message.append('! برای ردیف مقدار عددی (1, 2, 3 و ...) وارد کنید')
            self.ent_row.delete(0, END)
        
        finally :
            if len(error_message) == 2 :
                error_item = f'{error_message[0]}\n{error_message[1]}'
                messagebox.showerror('خظا', error_item)
            elif len(error_message) == 1 :
                error_item = f'{error_message[0]}'
                messagebox.showerror('خظا', error_item)
            else :
                self.title = self.ent_title.get()
                self.contest = self.ent_contest.get()
                self.column = int(self.ent_column.get())
                self.row = int(self.ent_row.get())
                self.setting_status = FALSE
                self.main_gui(self.frame_setting)

    def select_random(self):
        
        while TRUE:
            random_number = random.randrange(self.row, self.ws.max_row)
            winner = self.ws.cell(row= random_number, column= self.column).value
            if not winner in self.winners_list :
                break

        self.winners_list.append(winner)
        return(winner, self.winners_list.index(winner))
    
    def select_winner(self, event):

        x, i = self.select_random()
        row = i + 1
        name = x
        self.lbox_winners.insert(i, f'{row} -  {name}')
        if  i % 2 != 0 :
            self.lbox_winners.itemconfig(i, bg= '#ffca95')
        else :
            pass
        self.lbox_winners.see(END)

    def save_winners_list(self, event):
        wb = workbook.Workbook(write_only= TRUE)
        ws = wb.create_sheet(f'لیست اسامی برندگان {self.contest}')
        ws = wb.active
        winners = []
        for name in self.winners_list:
            item = []
            item.append(name)
            winners.append(item)
        for item in winners :
            ws.append(item)
        
        wb.save(os.path.join(os.path.expanduser('~'),'Desktop', f'اسامی قرعه کشی {self.title}.xlsx'))

        self.btn_save.config(text= 'لیست در دسکتاپ ذخیره شد')

    def open_file(self):
        try:
            file = filedialog.askopenfilename(filetypes= [('Excel Files', '*.xlsx')])
            self.wb = load_workbook(file)
            self.ws = self.wb.active
            self.contest_gui()
        except:
            pass

    def exit_app(self, event):
        if messagebox.askyesno('', 'آیا از برنامه خارج می شوید ؟'):
            root.destroy()
        else:
            pass


def main():
    
    global root

    root = Tk()

    #AppOpenInCenterOfScreen
    app_width = 360
    app_height = 640
    root.resizable(width= FALSE, height= FALSE)
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int((screen_width/2)-(app_width/2))
    y = int((screen_height/2)-(app_height/2))
    root.geometry(f'{app_width}x{app_height}+{x}+{y}')
    root.title('ابزار قرعه کشی')
    root.iconbitmap('ContestTool.ico')
    window = MainWindow(root)

    root.mainloop()    

if __name__ == '__main__' :
    main()